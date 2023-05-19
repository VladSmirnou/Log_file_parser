import datetime
import pathlib
import re

from openpyxl import load_workbook


class Main:
    num_pat = r'-?\d{1,}(?:\.\d{1,})?'
    bad_data_pat = re.compile(r"""(?P<stratname><[\w.-]+?>)
                            [\w\s]+:\s+(?P<depth>{0})%\s+
                            R:\s+(?P<rollback>{0})%\s+
                            d:\s+(?P<delta>{0})%""".format(num_pat), re.X)
    time_coin_price_pat = rf'(\d{{2}}:\d{{2}}:\d{{2}})[\w\s]+-(\w+)\s+Ask:({num_pat})'

    def __init__(self, user_inp):
        try:
            log_file_name, excel_file_name = [file_name.strip() for file_name in user_inp.split(',')]
        except:
            print(f'\nSomething wrong with your input. You typed -> [ {user_inp} ].\n')
            quit()

        self.log_file = self.get_log_file(log_file_name)
        self.excel_file_name = excel_file_name + '.xlsx'
        self.excel_file = self.get_excel_file()
        self.sheet_obj = self.excel_file.active
        
    def get_log_file(self, log_file_name):
        if log_file_name:
            log_file_name += '.log'
            try:
                log_file = open(log_file_name, 'r')
                return log_file
            except:
                print(f'\nLOG file with this name -> [ {log_file_name} ] doesn\'t exist in this directory\n')
        else:
            current_directory = pathlib.Path('.')
            for file in current_directory.iterdir():
                # if f'{datetime.date.today()}' in file.name:  # for the real use
                if '2023-03-26' in file.name:
                    print(f'\nFile found: {file.name}.\n')
                    log_file = open(file.name, 'r')
                    return log_file
                
            print(f'\nThere is no LOG file with today\'s date in this directory.\n')

        quit()

    def get_excel_file(self):
        try:
            return load_workbook(filename=self.excel_file_name)
        except:
            print(f'\nExcel file with this name -> [ {self.excel_file_name} ] doesn\'t exist in this directory.\n')
            quit()

    def get_excel_date_time(self):
        date = list(i.value.strftime('%Y-%m-%d') for i in self.sheet_obj['A'][1:])
        time = list(i.value.strftime('%H:%M:%S') for i in self.sheet_obj['B'][1:])
        return list(zip(date, time))

    def parse_signal(self, signal):
        left, right = signal.split(';', 1)
        time, coin, price = [g for g in re.search(self.time_coin_price_pat, left).groups()]
        not_ordered_data = re.search(self.bad_data_pat, right)
        ordered_data = re.findall(rf':\s({self.num_pat})\s', left)
        prep_list = [
            datetime.date.fromisoformat(self.date_from_log),
            datetime.time.fromisoformat(time),
            coin, 
            not_ordered_data['stratname'],
            price,
            not_ordered_data['depth'],
            not_ordered_data['rollback'],
            not_ordered_data['delta'],
            *ordered_data,
        ]
        return prep_list[:4] + [num.replace('.', ',') if '.' in num else num for num in prep_list[4:]]

    def parse_pump(self, pump):
        rem_prefix = pump[re.search(r'PumpQ=-?\d+\s', pump).end():]
        all_nums = re.findall(
            rf'[=\s]({self.num_pat})%?',
            ''.join(re.split(r'\s+sellX2.+?SellProb=.+?\s+', rem_prefix, 1))
        )
        return [num.replace('.', ',') if '.' in num else num for num in all_nums]

    def parse_emaf(self, emaf):
        all_nums = re.findall(rf'=\s({self.num_pat})%', emaf)
        return [num.replace('.', ',') if '.' in num else num for num in all_nums]

    def write_excel(self, complete_row):
        self.next_row = self.sheet_obj.max_row + 1
        for col_idx, value in enumerate(complete_row, 1):
            self.sheet_obj.cell(row=self.next_row, column=col_idx, value=value)

    def data_validation(self):
        date_time_tp_excel = self.get_excel_date_time()

        self.date_from_log = re.search(r'\d{4}-\d{2}-\d{2}', self.log_file.name).group(0)

        rows_before_upd = self.sheet_obj.max_row

        row_count = 0
        for idx, line in enumerate(list_of_lines:= self.log_file.read().splitlines()):
            if time_from_log:= re.match(r'(\d{2}:\d{2}:\d{2})\s+Signal', line):
                if not (self.date_from_log, time_from_log.group(1)) in date_time_tp_excel:
                    row_count += 1
                    signal = list_of_lines[idx]
                    pump = list_of_lines[idx + 1]
                    emaf = list_of_lines[idx + 2]

                    signal_portion = self.parse_signal(signal)
                    pump_portion = self.parse_pump(pump)
                    emaf_portion = self.parse_emaf(emaf)

                    complete_row = signal_portion + pump_portion + emaf_portion

                    self.write_excel(complete_row)
                    
        if not row_count:
            print(f'\nThere is no new data in [ {self.log_file.name} ] file.\n')
        else:
            self.excel_file.save(self.excel_file_name)
            print(
                f'\n[ SUCCESS ] rows added to excel: {row_count}. Last row before updating: {rows_before_upd}. Last row now: {self.next_row}.\n'
            )

        self.log_file.close()
        self.excel_file.close()


if __name__ == '__main__':
    print('Type LOG file name without \'.log\' part that you wanna parse and excel file name, that you wanna insert data to. Separate them by comma.')
    print('You can omit LOG file name, then it will try to find the first file with today\'s date in this directory.')
    print('You cannot omit excel file name.')
    print('-' * 10)
    print('Example with LOG file name: "LOG_2023-23-23, log_excel".')
    print('Example without LOG file name: ", log_excel".')
    print('-' * 10)
    obj = Main(input('Type here: '))
    obj.data_validation()
